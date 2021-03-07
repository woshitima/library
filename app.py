from flask import Flask, render_template, request
from openpyxl import load_workbook
from sqlalchemy.orm.session import sessionmaker
from sqlalchemy.orm import scoped_session
from database import engine, Book

app = Flask(__name__)
db = scoped_session(sessionmaker(bind=engine))

@app.route("/")
def homepage():
    return render_template("index.html")


@app.route("/us/")
def info():
    return render_template("aboutus.html")


@app.route("/form/")
def form():
    return render_template("form.html")


@app.route("/books/")
def books():
    if 'key_word' in request.args:
        key_word = request.args.get("key_word")
        session = sessionmaker(engine)()
        books = session.execute(f"""
            SELECT *
            FROM "Book"
            WHERE name LIKE '%{key_word}%'
                    OR author LIKE '%{key_word}%';
            """)
        session.commit()
    else:
        books = db.execute("""SELECT * FROM "Book";""")
        db.commit()

    return render_template("books_table.html", object_list=books)


@app.route("/authors/")
def authors():
    with engine.connect() as con:
        authors = con.execute('SELECT DISTINCT author FROM "Book";')

    return render_template(
        "authors.html", authors=list(authors)
    )

@app.route("/db/authors/")
def db_authors():
    with engine.connect() as con:
        authors = con.execute('SELECT DISTINCT author FROM "Book";')
 
    return render_template(
        "database_authors.html", authors=authors
    )


@app.route("/add/", methods=["POST"])
def add():
    f = request.form
    book = f["book"]
    author = f["author"]
    url = f["url"]    
    
    ids = db.execute('SELECT id FROM "Book" ORDER BY id DESC;')
    max_id = ids.first().id
    c_id = max_id + 1

    db.execute(f'''
        INSERT INTO "Book" (id, name, author, image)
        VALUES ({c_id}, '{book}', '{author}', '{url}');''')
    db.commit()

    message = "Form Received!"
    
    return render_template("index.html", message = message)


@app.route("/book/<num>/")
def book(num):
    excel = load_workbook("tales.xlsx")
    page = excel["Sheet"]
    object_list = [[tale.value, tale.offset(column=1).value, tale.offset(column=2).value] for tale in page["A"][1:]]
    obj = object_list[int(num)]
    obj.append(num)
    return render_template("book.html", obj=obj)


@app.route("/db/book/<id>/")
def db_book(id):
    obj = db.execute(f'SELECT * FROM "Book" WHERE id = {id};').first()
    return render_template("database_book.html", obj=obj)


@app.route("/book/<num>/save/", methods=["POST"])
def book_save(num):
    num = int(num)
    excel_file = load_workbook("tales.xlsx")
    page = excel_file["Sheet"]
    form = request.form
    page[f"A{num}"] = form["tale"]
    page[f"B{num}"] = form["author"]
    page[f"C{num}"] = form["image"]
    excel_file.save("tales.xlsx")
    return "Saved!"


@app.route("/<int:id>/", methods=["GET", "POST"])
def db_book_update(id):
    message = ''
    if request.method == "POST":
        name = request.form.get("tale")
        author = request.form.get("author")
        image = request.form.get("image")
        db.execute(f'''
            UPDATE "Book"
            SET 
                name='{name}',
                author='{author}',
                image='{image}'
            WHERE id={id};
        ''')
        db.commit()
        message = "Changes Saved!"

    book_object = db.execute(f'SELECT * FROM "Book" WHERE id={id};').first()
    return render_template("database_book_update.html", book_object=book_object, message=message)